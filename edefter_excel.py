from __future__ import annotations

import re
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Callable

try:
    from bs4 import BeautifulSoup
    BS4_AVAILABLE = True
except Exception as _bs4_exc:
    BeautifulSoup = None
    BS4_AVAILABLE = False
    BS4_IMPORT_ERROR = str(_bs4_exc)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    OPENPYXL_AVAILABLE = True
except Exception as _openpyxl_exc:
    Workbook = None
    Alignment = Font = PatternFill = None
    OPENPYXL_AVAILABLE = False
    OPENPYXL_IMPORT_ERROR = str(_openpyxl_exc)

try:
    from pypdf import PdfReader
    PYPDF_AVAILABLE = True
except Exception as _pypdf_exc:
    PdfReader = None
    PYPDF_AVAILABLE = False
    PYPDF_IMPORT_ERROR = str(_pypdf_exc)


HEADERS = [
    "Hesap Kodu",
    "Evrak Tarihi",
    "Evrak No",
    "B.T.",
    "Vergi/TC No",
    "Açıklama",
    "Borçlu",
    "Alacaklı",
    "Belge Türü Açıklaması (B.Türü 8 ise)",
    "Ödeme Şekli",
]


HtmlBuilder = Callable[[Path, Path | None], Path]


def _ensure_conversion_dependencies() -> None:
    missing: list[str] = []
    if not BS4_AVAILABLE:
        missing.append(f"beautifulsoup4 ({BS4_IMPORT_ERROR})")
    if not OPENPYXL_AVAILABLE:
        missing.append(f"openpyxl ({OPENPYXL_IMPORT_ERROR})")
    if missing:
        raise RuntimeError("Excel aktarımı için eksik bağımlılıklar: " + ", ".join(missing))


def _clean_text(value: str) -> str:
    return " ".join(value.replace("\xa0", " ").split())


def _parse_turkish_amount(value: str) -> float:
    return float(value.replace(".", "").replace(",", "."))


def _parse_doc_line(text: str) -> tuple[str, str, str]:
    match = re.search(
        r"(?P<label>.+?)\s+No\s*:\s*(?P<number>\S+)(?:\s+(?P=label)\s+Tarihi\s*:\s*(?P<date>\d{2}/\d{2}/\d{4}))?",
        text,
    )
    if not match:
        return "", "", ""
    return (
        _clean_text(match.group("label")),
        match.group("number"),
        match.group("date") or "",
    )


def _parse_footer_text(text: str) -> tuple[str, str]:
    lines = [_clean_text(line) for line in text.splitlines() if _clean_text(line)]
    section = lines[0] if lines else ""
    fis_no = ""
    for line in lines:
        match = re.search(r"Muhasebe Fiş No\s*:\s*(\S+)", line)
        if match:
            fis_no = match.group(1)
            break
    return section, fis_no


def _extract_pdf_text(pdf_path: Path, max_pages: int = 4) -> str:
    if not PYPDF_AVAILABLE:
        raise RuntimeError(f"PDF desteği için pypdf eksik: {PYPDF_IMPORT_ERROR}")

    reader = PdfReader(str(pdf_path))
    pages = reader.pages[:max_pages]
    return "\n".join((page.extract_text() or "") for page in pages)


def _find_paths_in_pdf_text(pdf_path: Path) -> list[Path]:
    text = _extract_pdf_text(pdf_path)
    matches = re.findall(r"[A-Za-z]:\\[^\r\n]+?\.(?:xml|html)", text, flags=re.IGNORECASE)
    candidates: list[Path] = []

    for match in matches:
        candidate = Path(match.strip())
        if candidate.suffix.lower() not in {".xml", ".html"}:
            continue
        if candidate not in candidates:
            candidates.append(candidate)

        file_name_match = re.search(r"([^\\/:]+\.(?:xml|html))$", match, flags=re.IGNORECASE)
        if file_name_match:
            fallback = (pdf_path.parent / file_name_match.group(1)).resolve()
            if fallback not in candidates:
                candidates.append(fallback)

    return candidates


def _find_sibling_sources(input_path: Path) -> list[Path]:
    folder = input_path.parent
    candidates: list[Path] = []

    exact_html = input_path.with_suffix(".html")
    exact_xml = input_path.with_suffix(".xml")
    for candidate in (exact_html, exact_xml):
        if candidate.exists() and candidate not in candidates:
            candidates.append(candidate)

    patterns = ["*-Y-*.html", "*-Y-*.xml", "*.html", "*.xml"]
    for pattern in patterns:
        for candidate in sorted(folder.glob(pattern)):
            resolved = candidate.resolve()
            if resolved not in candidates:
                candidates.append(resolved)
    return candidates


def inspect_yevmiye_input(input_path: str | Path) -> dict[str, str | Path | None]:
    path = Path(input_path).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"Dosya bulunamadı: {path}")

    suffix = path.suffix.lower()
    if suffix == ".xml":
        return {
            "input_type": "xml",
            "source_type": "xml",
            "resolved_source": path,
            "message": "XML dosyası doğrudan kullanılacak.",
        }

    if suffix == ".html":
        return {
            "input_type": "html",
            "source_type": "html",
            "resolved_source": path,
            "message": "HTML dosyası doğrudan kullanılacak.",
        }

    if suffix == ".pdf":
        errors: list[str] = []
        if PYPDF_AVAILABLE:
            try:
                for candidate in _find_paths_in_pdf_text(path):
                    if candidate.exists():
                        return {
                            "input_type": "pdf",
                            "source_type": candidate.suffix.lower().lstrip("."),
                            "resolved_source": candidate,
                            "message": "PDF içindeki yevmiye kaynak dosyası bulundu.",
                        }
            except Exception as exc:
                errors.append(str(exc))

        for candidate in _find_sibling_sources(path):
            if candidate.exists():
                return {
                    "input_type": "pdf",
                    "source_type": candidate.suffix.lower().lstrip("."),
                    "resolved_source": candidate,
                    "message": "PDF ile aynı klasörde eşleşen yevmiye kaynağı bulundu.",
                }

        detail = f" Detay: {' | '.join(errors)}" if errors else ""
        raise FileNotFoundError(
            "PDF içinden kullanılabilir XML/HTML kaynağı bulunamadı. "
            "Aynı klasörde yevmiye XML/HTML dosyası olmalı veya PDF raporunun içinde kaynak yol geçmeli."
            + detail
        )

    raise ValueError("Desteklenmeyen dosya türü. PDF, XML veya HTML seçiniz.")


def resolve_input_to_html(input_path: str | Path, html_builder: HtmlBuilder) -> tuple[Path, dict[str, str | Path | None]]:
    info = inspect_yevmiye_input(input_path)
    resolved_source = info["resolved_source"]
    if not isinstance(resolved_source, Path):
        raise RuntimeError("Çözülen kaynak dosya geçersiz.")

    if resolved_source.suffix.lower() == ".html":
        return resolved_source, info

    if resolved_source.suffix.lower() != ".xml":
        raise ValueError("Yevmiye aktarımı için çözülen kaynak XML veya HTML olmalı.")

    temp_dir = Path(tempfile.gettempdir()) / "edefter_bot_excel"
    temp_dir.mkdir(parents=True, exist_ok=True)
    html_path = temp_dir / f"{resolved_source.stem}.html"
    built_html = html_builder(resolved_source, html_path)
    return Path(built_html).resolve(), info


def parse_entries(html_path: str | Path) -> list[list[object]]:
    _ensure_conversion_dependencies()

    soup = BeautifulSoup(Path(html_path).read_text(encoding="utf-8"), "lxml")
    body = soup.body
    if body is None:
        raise ValueError("HTML içinde body etiketi bulunamadı.")

    children = [child for child in body.children if getattr(child, "name", None)]
    rows: list[list[object]] = []
    index = 0

    while index < len(children):
        node = children[index]
        classes = node.get("class") or []
        if node.name != "table" or "entryHeaderHeader" not in classes:
            index += 1
            continue

        header_cells = [_clean_text(td.get_text(" ", strip=True)) for td in node.find_all("td")]
        voucher_date_text = ""
        if len(header_cells) > 1:
            voucher_match = re.search(r"\[\s*(\d{2}/\d{2}/\d{4})\s*\]", header_cells[1])
            if voucher_match:
                voucher_date_text = voucher_match.group(1)

        index += 1
        detail_tables = []
        footer_text = ""
        while index < len(children):
            current = children[index]
            current_classes = current.get("class") or []
            if current.name == "table" and "entryDetail" in current_classes:
                detail_tables.append(current)
                index += 1
                continue
            if current.name == "div" and "entryHeaderFooter" in current_classes:
                footer_text = current.get_text("\n", strip=True)
                index += 1
                break
            index += 1

        _section_title, fis_no = _parse_footer_text(footer_text)

        for table in detail_tables:
            tr_nodes = table.find_all("tr")
            if not tr_nodes:
                continue

            first_row_cells = [_clean_text(td.get_text(" ", strip=True)) for td in tr_nodes[0].find_all("td")]
            amount_cells = [cell for cell in first_row_cells if re.fullmatch(r"[\d.]+,\d{2}", cell)]
            amount = _parse_turkish_amount(amount_cells[-1]) if amount_cells else 0.0
            debit = amount if len(first_row_cells) >= 3 and first_row_cells[-2] else 0.0
            credit = amount if len(first_row_cells) >= 4 and first_row_cells[-1] else 0.0

            detailed_code = ""
            description = ""
            payment_method = ""
            document_label = ""
            document_number = ""
            document_date = voucher_date_text
            belge_turu = ""
            belge_turu_aciklamasi = ""

            for tr_node in tr_nodes[1:]:
                line_text = _clean_text(tr_node.get_text(" ", strip=True))
                if not line_text:
                    continue
                if re.match(r"^\d", line_text):
                    detailed_code = line_text.split()[0]
                    continue
                if line_text.startswith("-"):
                    description = line_text.lstrip("-").strip()
                    continue
                if "Ödeme Şekli :" in line_text:
                    payment_method = _clean_text(line_text.split("Ödeme Şekli :", 1)[1])
                doc_label, doc_number, doc_date = _parse_doc_line(line_text)
                if doc_number:
                    document_label = doc_label
                    document_number = doc_number
                    document_date = doc_date or document_date

            if document_label == "Fatura":
                belge_turu = "1"
            elif document_label == "Serbest Meslek Makbuzu":
                belge_turu = "8"
                belge_turu_aciklamasi = document_label
            elif document_label:
                belge_turu_aciklamasi = document_label

            if not document_date:
                raise ValueError("Yevmiye satırı için evrak tarihi çözümlenemedi.")

            effective_doc_no = document_number or fis_no
            effective_date = datetime.strptime(document_date, "%d/%m/%Y")

            rows.append(
                [
                    detailed_code,
                    effective_date,
                    effective_doc_no,
                    belge_turu,
                    "",
                    description,
                    debit,
                    credit,
                    belge_turu_aciklamasi,
                    payment_method,
                ]
            )

    if not rows:
        raise ValueError("HTML içinde aktarılabilir yevmiye satırı bulunamadı.")

    return rows


def build_workbook(rows: list[list[object]]):
    _ensure_conversion_dependencies()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Yevmiye Aktarım"
    worksheet.append(HEADERS)

    for row in rows:
        worksheet.append(row)

    header_fill = PatternFill(fill_type="solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)
    centered = Alignment(horizontal="center", vertical="center")

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = centered

    widths = {
        "A": 18,
        "B": 14,
        "C": 18,
        "D": 8,
        "E": 16,
        "F": 45,
        "G": 14,
        "H": 14,
        "I": 30,
        "J": 20,
    }
    for column_name, width in widths.items():
        worksheet.column_dimensions[column_name].width = width

    for row_index in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row_index, column=2).number_format = "dd.mm.yyyy"
        worksheet.cell(row=row_index, column=7).number_format = "0.00"
        worksheet.cell(row=row_index, column=8).number_format = "0.00"

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    return workbook


def export_yevmiye_to_excel(input_path: str | Path, output_path: str | Path, html_builder: HtmlBuilder) -> dict[str, object]:
    html_path, info = resolve_input_to_html(input_path, html_builder)
    rows = parse_entries(html_path)
    workbook = build_workbook(rows)

    target = Path(output_path).expanduser().resolve()
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)

    return {
        "row_count": len(rows),
        "output_path": target,
        "html_path": html_path,
        "source_type": info["source_type"],
        "resolved_source": info["resolved_source"],
        "message": info["message"],
    }